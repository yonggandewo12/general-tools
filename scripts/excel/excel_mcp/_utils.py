"""Shared utilities for Excel MCP modules.

集中放置跨模块复用的能力：工作簿资源管理、单元格引用解析、列字母转换、
路径安全、JSON 序列化。统一这些底层逻辑可消除各模块重复代码，确保资源
释放一致、错误处理一致、跨平台行为一致。
"""

from __future__ import annotations

import os
import re
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import NotFoundError, ValidationError


# ────────────────────────────── 路径 ──────────────────────────────

def resolve_path(filepath: str) -> str:
    """规范化文件路径，跨平台兼容。

    - 接受绝对/相对路径，统一返回绝对路径字符串
    - 拒绝包含 NUL 字节的路径（安全）
    - 不做沙箱限制（stdio 模式下由调用方控制路径）
    """
    if not filepath or "\x00" in filepath:
        raise ValidationError(f"Invalid filepath: {filepath!r}")
    return str(Path(filepath).expanduser().resolve())


# ────────────────────────────── 工作簿资源管理 ──────────────────────────────

@contextmanager
def open_workbook(filepath: str, *, read_only: bool = False, data_only: bool = False) -> Iterator[Any]:
    """以 context manager 打开工作簿，确保最终 close()。

    用法:
        with open_workbook(path) as wb:
            ws = wb[sheet]
            ...
    不负责保存——需要写入时用 save_workbook() 或在 with 块内 wb.save()。
    """
    path = resolve_path(filepath)
    if not Path(path).exists():
        raise NotFoundError(f"File not found: {path}")
    wb = load_workbook(path, read_only=read_only, data_only=data_only)
    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception:
            pass


@contextmanager
def edit_workbook(filepath: str) -> Iterator[Any]:
    """打开工作簿用于编辑，退出时自动保存。

    用法:
        with edit_workbook(path) as wb:
            wb[sheet]["A1"] = 1
    自动 save + close，即使异常也会 close（不保存异常结果）。
    """
    path = resolve_path(filepath)
    if not Path(path).exists():
        raise NotFoundError(f"File not found: {path}")
    wb = load_workbook(path)
    try:
        yield wb
        wb.save(path)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def get_or_create_workbook(filepath: str) -> Any:
    """获取已有工作簿或新建一个（不保存，由调用方处理）。"""
    path = resolve_path(filepath)
    if Path(path).exists():
        return load_workbook(path)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb["Sheet"].title = "Sheet1"
    return wb


def require_sheet(wb: Any, sheet_name: str) -> Worksheet:
    """取出工作表，不存在则抛 SheetError。"""
    from .exceptions import SheetError
    if sheet_name not in wb.sheetnames:
        raise SheetError(f"Sheet {sheet_name!r} not found")
    return wb[sheet_name]


# ────────────────────────────── 单元格引用解析 ──────────────────────────────

_CELL_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
_COL_RE = re.compile(r"^[A-Z]+$")


def validate_cell_ref(cell_ref: str) -> bool:
    """校验单元格引用格式（如 'A1', 'BC123'）。"""
    if not cell_ref:
        return False
    return bool(_CELL_RE.match(cell_ref.upper()))


def parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """解析单个单元格引用为 (row, col)，均为 1-based。

    Raises ValueError if invalid.
    """
    m = _CELL_RE.match(cell_ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {cell_ref!r}")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    return row, col


def parse_cell_range(
    cell_ref: str, end_ref: str | None = None
) -> tuple[int, int, int | None, int | None]:
    """解析范围为 (start_row, start_col, end_row, end_col)。

    - 单个引用: end_row/end_col 为 None
    - 含 ':' 的引用如 'A1:B2' 自动拆分
    - end_ref 显式提供时优先
    """
    if end_ref is None and ":" in cell_ref:
        start_cell, end_cell = cell_ref.split(":", 1)
        start_row, start_col = parse_cell_ref(start_cell)
        end_row, end_col = parse_cell_ref(end_cell)
        return start_row, start_col, end_row, end_col

    start_row, start_col = parse_cell_ref(cell_ref)
    if end_ref:
        end_row, end_col = parse_cell_ref(end_ref)
        return start_row, start_col, end_row, end_col
    return start_row, start_col, None, None


def range_to_str(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def col_letter_to_index(letters: str) -> int:
    return column_index_from_string(letters.upper())


def col_index_to_letter(idx: int) -> str:
    return get_column_letter(idx)


# ────────────────────────────── JSON 序列化 ──────────────────────────────

def to_json_safe(value: Any) -> Any:
    """把 openpyxl 值转为 JSON 安全类型。

    处理 datetime/date/time/Decimal/Color 等 json.dumps default=str 兜不住的边界。
    """
    import datetime as _dt
    from decimal import Decimal

    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, _dt.datetime):
        return value.isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, _dt.time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    # openpyxl Color / 其他对象
    return str(value)
