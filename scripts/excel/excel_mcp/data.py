"""数据读写：write_data / read_excel_range / read_excel_range_with_metadata。"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.utils import get_column_letter

from ._utils import edit_workbook, open_workbook, parse_cell_range, require_sheet, to_json_safe
from .cell_validation import get_data_validation_for_cell
from .exceptions import DataError, ValidationError

logger = logging.getLogger(__name__)


def _resolve_range_bounds(ws, start_cell: str, end_cell: str | None) -> tuple[int, int, int, int]:
    """解析读取范围，end_cell 缺省时自动扩展到工作表数据边界。"""
    # 支持 'A1:B2' 形式的 start_cell
    if end_cell is None and ":" in start_cell:
        start_cell, end_cell = start_cell.split(":", 1)
    sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
    if er is None or ec is None:
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            er, ec = sr, sc
        else:
            # 默认从 start_cell 开始到工作表末尾
            er, ec = ws.max_row, ws.max_column
            if start_cell.upper() == "A1":
                sr, sc = ws.min_row, ws.min_column
    return sr, sc, er, ec


def read_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False,
) -> list[list[Any]]:
    """读取范围内的数据，返回二维列表（按行）。"""
    try:
        with open_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            sr, sc, er, ec = _resolve_range_bounds(ws, start_cell, end_cell)
            if sr > ws.max_row or sc > ws.max_column:
                logger.warning(
                    "Start cell %s outside data boundary %s:%s; no data read.",
                    start_cell,
                    get_column_letter(ws.min_column) + str(ws.min_row),
                    get_column_letter(ws.max_column) + str(ws.max_row),
                )
                return []
            data: list[list[Any]] = []
            for r in range(sr, er + 1):
                row_data = [to_json_safe(ws.cell(row=r, column=c).value) for c in range(sc, ec + 1)]
                if any(v is not None for v in row_data):
                    data.append(row_data)
            if preview_only:
                return data[:10]
            return data
    except (DataError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to read Excel range: %s", e)
        raise DataError(str(e)) from e


def write_data(
    filepath: str,
    sheet_name: str | None,
    data: list[list[Any]] | None,
    start_cell: str = "A1",
) -> dict[str, str]:
    """将二维列表写入工作表，起始单元格默认 A1。工作表不存在则创建。"""
    try:
        if not data:
            raise DataError("No data provided to write")
        with edit_workbook(filepath) as wb:
            if not sheet_name:
                active = wb.active
                if active is None:
                    raise DataError("No active sheet found in workbook")
                sheet_name = active.title
            elif sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            sr, sc = parse_cell_range(start_cell)[0:2]
            for i, row in enumerate(data):
                if not isinstance(row, list):
                    raise DataError(f"Row {i} is not a list: {type(row).__name__}")
                for j, val in enumerate(row):
                    ws.cell(row=sr + i, column=sc + j, value=val)
        return {"message": f"Data written to {sheet_name}", "active_sheet": sheet_name}
    except (DataError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to write data: %s", e)
        raise DataError(str(e)) from e


def read_excel_range_with_metadata(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    include_validation: bool = True,
    preview_only: bool = False,
) -> dict[str, Any]:
    """读取范围数据并附带每个单元格的元数据（地址、行、列、验证规则）。"""
    try:
        with open_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            sr, sc, er, ec = _resolve_range_bounds(ws, start_cell, end_cell)
            if sr > ws.max_row or sc > ws.max_column:
                rng = f"{start_cell}:{end_cell}" if end_cell else f"{start_cell}:"
                return {"range": rng, "sheet_name": sheet_name, "cells": []}
            if preview_only:
                er = min(er, sr + 9)
            rng = f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}"
            cells: list[dict[str, Any]] = []
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    cell = ws.cell(row=r, column=c)
                    addr = f"{get_column_letter(c)}{r}"
                    cell_data: dict[str, Any] = {
                        "address": addr,
                        "value": to_json_safe(cell.value),
                        "row": r,
                        "column": c,
                    }
                    if include_validation:
                        v = get_data_validation_for_cell(ws, addr)
                        cell_data["validation"] = v if v else {"has_validation": False}
                    cells.append(cell_data)
            return {"range": rng, "sheet_name": sheet_name, "cells": cells}
    except (DataError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to read Excel range with metadata: %s", e)
        raise DataError(str(e)) from e
