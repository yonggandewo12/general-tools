"""透视汇总表（基于源数据手动聚合，输出为新工作表 + 原生 Table 样式）。"""

from __future__ import annotations

import logging
import uuid
from typing import Any

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ._utils import edit_workbook, parse_cell_range, require_sheet
from .data import read_excel_range
from .exceptions import PivotError, ValidationError

logger = logging.getLogger(__name__)

_VALID_AGG = {"sum", "average", "count", "min", "max"}
_AGG_SUFFIXES = (" (sum)", " (average)", " (count)", " (min)", " (max)")


def _clean_field(field: Any) -> str:
    f = str(field).strip()
    low = f.lower()
    for suf in _AGG_SUFFIXES:
        if low.endswith(suf):
            return f[: -len(suf)]
    return f


def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: list[str],
    values: list[str],
    columns: list[str] | None = None,
    agg_func: str = "sum",
) -> dict[str, Any]:
    try:
        if ":" not in data_range:
            raise ValidationError("Data range must be in format 'A1:B2'")
        start_cell, end_cell = data_range.split(":", 1)
        sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
        if er is None or ec is None:
            raise ValidationError("Invalid data range format: missing end coordinates")

        if agg_func.lower() not in _VALID_AGG:
            raise ValidationError(f"Invalid aggregation function. Must be one of: {', '.join(sorted(_VALID_AGG))}")

        raw = read_excel_range(filepath, sheet_name, start_cell, end_cell)
        if not raw or len(raw) < 2:
            raise PivotError("Source data must have a header row and at least one data row.")
        headers = [str(h) for h in raw[0]]
        records = [dict(zip(headers, r)) for r in raw[1:]]
        if not records:
            raise PivotError("No data rows found after header.")

        available = {str(h).lower() for h in headers}
        for field_list, label in [(rows, "row"), (values, "value")]:
            for f in field_list:
                if _clean_field(f).lower() not in available:
                    raise ValidationError(
                        f"Invalid {label} field {f!r}. Available: {', '.join(sorted(headers))}"
                    )
        if columns:
            for f in columns:
                if _clean_field(f).lower() not in available:
                    raise ValidationError(f"Invalid column field {f!r}. Available: {', '.join(sorted(headers))}")

        cleaned_rows = [_clean_field(f) for f in rows]
        cleaned_values = [_clean_field(f) for f in values]

        with edit_workbook(filepath) as wb:
            require_sheet(wb, sheet_name)
            pivot_name = f"{sheet_name}_pivot"
            if pivot_name in wb.sheetnames:
                wb.remove(wb[pivot_name])
            pws = wb.create_sheet(pivot_name)

            # 表头
            col = 1
            for f in cleaned_rows:
                pws.cell(row=1, column=col, value=f).font = Font(bold=True)
                col += 1
            for f in cleaned_values:
                pws.cell(row=1, column=col, value=f"{f} ({agg_func})").font = Font(bold=True)
                col += 1

            # 行字段唯一值组合
            field_values: dict[str, list[str]] = {}
            for f in cleaned_rows:
                field_values[f] = sorted({str(rec.get(f, "")) for rec in records})
            combos = _combinations(field_values)

            total_rows = len(combos) + 1
            total_cols = len(cleaned_rows) + len(cleaned_values)

            r = 2
            for combo in combos:
                c = 1
                for f in cleaned_rows:
                    pws.cell(row=r, column=c, value=combo[f])
                    c += 1
                filtered = [rec for rec in records if all(str(rec.get(k, "")) == v for k, v in combo.items())]
                for vf in cleaned_values:
                    try:
                        pws.cell(row=r, column=c, value=_aggregate(filtered, vf, agg_func))
                    except Exception as e:
                        raise PivotError(f"Failed to aggregate field {vf!r}: {e}") from e
                    c += 1
                r += 1

            # 套用 Table 样式
            try:
                ref = f"A1:{get_column_letter(total_cols)}{total_rows}"
                tbl = Table(displayName=f"PivotTable_{uuid.uuid4().hex[:8]}", ref=ref)
                tbl.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True,
                )
                pws.add_table(tbl)
            except Exception as e:
                raise PivotError(f"Failed to apply pivot table style: {e}") from e

        return {
            "message": "Summary table created successfully",
            "details": {
                "source_range": data_range,
                "pivot_sheet": pivot_name,
                "rows": cleaned_rows,
                "columns": columns or [],
                "values": cleaned_values,
                "aggregation": agg_func,
            },
        }
    except (ValidationError, PivotError):
        raise
    except Exception as e:
        logger.error("Failed to create pivot table: %s", e)
        raise PivotError(str(e)) from e


def _combinations(field_values: dict[str, list[str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = [{}]
    for field, vals in field_values.items():
        result = [{**combo, field: v} for combo in result for v in vals]
    return result


def _aggregate(records: list[dict[str, Any]], field: str, agg_func: str) -> float:
    nums = [r[field] for r in records if field in r and isinstance(r[field], (int, float))]
    if not nums:
        return 0
    af = agg_func.lower()
    if af == "sum":
        return sum(nums)
    if af == "average":
        return sum(nums) / len(nums)
    if af == "count":
        return len(nums)
    if af == "min":
        return min(nums)
    if af == "max":
        return max(nums)
    return sum(nums)
