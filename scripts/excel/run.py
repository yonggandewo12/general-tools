#!/usr/bin/env python3
"""Excel MCP 统一入口脚本。

被 src/excel-service.ts 通过子进程调用。协议：
  python run.py --action <name> --params '<json>'
  python run.py --list          # 列出所有 action
  python run.py --check         # 自检依赖
  python run.py --action <name> # params 从 stdin 读 JSON

输出（stdout）固定为单行 JSON：
  成功: {"success": true, "data": <result>}
  失败: {"success": false, "error": "...", "code": "...", "error_type": "..."}

日志走 stderr，不污染 stdout。
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import traceback
from pathlib import Path
from typing import Any, Callable

# 确保本脚本所在目录在 sys.path 最前，便于 `import excel_mcp`
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from excel_mcp import (  # noqa: E402
    calculations,
    chart,
    data,
    formatting,
    pivot,
    sheet,
    tables,
    validation,
    workbook,
)
from excel_mcp.cell_validation import get_all_validation_ranges  # noqa: E402
from excel_mcp.exceptions import ExcelMCPError, NotFoundError  # noqa: E402
from excel_mcp._utils import open_workbook, require_sheet, to_json_safe  # noqa: E402

logger = logging.getLogger("excel_mcp.run")


# ────────────────────────────── dispatch 注册 ──────────────────────────────
# 每个条目: action_name -> callable(**params)
# callable 直接调用对应模块函数，返回 dict/list/str。
# 新增工具只需在此注册一行 —— 可扩展性集中体现。

ACTIONS: dict[str, Callable[..., Any]] = {
    # 工作簿
    "create_workbook": workbook.create_workbook,
    "create_worksheet": workbook.create_sheet,
    "get_workbook_metadata": workbook.get_workbook_info,
    # 数据
    "write_data": data.write_data,
    "read_data": data.read_excel_range_with_metadata,
    # 公式
    "apply_formula": calculations.apply_formula,
    "validate_formula_syntax": validation.validate_formula_in_cell_operation,
    # 格式化
    "format_range": formatting.format_range,
    # 合并
    "merge_cells": sheet.merge_range,
    "unmerge_cells": sheet.unmerge_range,
    "get_merged_cells": sheet.get_merged_ranges,
    # 图表
    "create_chart": chart.create_chart_in_sheet,
    # 透视
    "create_pivot_table": pivot.create_pivot_table,
    # 表
    "create_table": tables.create_excel_table,
    # 工作表
    "copy_worksheet": sheet.copy_sheet,
    "delete_worksheet": sheet.delete_sheet,
    "rename_worksheet": sheet.rename_sheet,
    # 范围
    "copy_range": sheet.copy_range_operation,
    "delete_range": sheet.delete_range_operation,
    "validate_excel_range": validation.validate_range_in_sheet_operation,
    # 行列
    "insert_rows": sheet.insert_row,
    "insert_columns": sheet.insert_cols,
    "delete_sheet_rows": sheet.delete_rows,
    "delete_sheet_columns": sheet.delete_cols,
}


def _get_data_validation_info(filepath: str, sheet_name: str) -> dict[str, Any]:
    """获取工作表内所有数据验证规则。"""
    with open_workbook(filepath) as wb:
        ws = require_sheet(wb, sheet_name)
        validations = get_all_validation_ranges(ws)
    if not validations:
        return {"sheet_name": sheet_name, "validation_rules": []}
    return {"sheet_name": sheet_name, "validation_rules": validations}


ACTIONS["get_data_validation_info"] = _get_data_validation_info


# ────────────────────────────── 输出 ──────────────────────────────

def _emit(obj: dict[str, Any]) -> None:
    """单行 JSON 输出到 stdout。"""
    sys.stdout.write(json.dumps(obj, default=to_json_safe, ensure_ascii=False))
    sys.stdout.write("\n")
    sys.stdout.flush()


def _ok(result: Any) -> None:
    # 字符串结果包装为 message
    if isinstance(result, str):
        _emit({"success": True, "data": {"message": result}})
    else:
        _emit({"success": True, "data": result})


def _fail(err: BaseException) -> None:
    if isinstance(err, ExcelMCPError):
        _emit({"success": False, "error": str(err), "code": getattr(err, "code", "EXCEL_ERROR"), "error_type": err.__class__.__name__})
    else:
        _emit({"success": False, "error": str(err), "code": "INTERNAL_ERROR", "error_type": err.__class__.__name__})


# ────────────────────────────── 参数解析 ──────────────────────────────

def _load_params(args: argparse.Namespace) -> dict[str, Any]:
    if args.params is not None:
        if args.params.strip() == "":
            return {}
        return json.loads(args.params)
    # 从 stdin 读
    if not sys.stdin.isatty():
        raw = sys.stdin.read().strip()
        if raw:
            return json.loads(raw)
    return {}


def _setup_logging() -> None:
    level = logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )


def main() -> int:
    _setup_logging()
    parser = argparse.ArgumentParser(description="Excel MCP entry")
    parser.add_argument("--action", help="action name")
    parser.add_argument("--params", help="JSON params (default: read from stdin)")
    parser.add_argument("--list", action="store_true", help="list all actions")
    parser.add_argument("--check", action="store_true", help="self-check dependencies")
    args = parser.parse_args()

    if args.list:
        _emit({"success": True, "data": {"actions": sorted(ACTIONS.keys())}})
        return 0

    if args.check:
        try:
            import openpyxl  # noqa: F401
            _emit({"success": True, "data": {"openpyxl": openpyxl.__version__, "python": sys.version.split()[0]}})
            return 0
        except Exception as e:
            _emit({"success": False, "error": f"openpyxl not available: {e}", "code": "DEP_MISSING"})
            return 1

    if not args.action:
        _emit({"success": False, "error": "No --action provided", "code": "MISSING_ACTION"})
        return 1

    fn = ACTIONS.get(args.action)
    if fn is None:
        _emit({"success": False, "error": f"Unknown action: {args.action}", "code": "UNKNOWN_ACTION", "available": sorted(ACTIONS.keys())})
        return 1

    try:
        params = _load_params(args)
    except json.JSONDecodeError as e:
        _emit({"success": False, "error": f"Invalid JSON params: {e}", "code": "BAD_PARAMS"})
        return 1

    try:
        result = fn(**params)
        _ok(result)
        return 0
    except ExcelMCPError as e:
        logger.warning("action %s failed: %s", args.action, e)
        _fail(e)
        return 2
    except Exception as e:
        logger.error("action %s crashed: %s\n%s", args.action, e, traceback.format_exc())
        _fail(e)
        return 3


if __name__ == "__main__":
    sys.exit(main())
