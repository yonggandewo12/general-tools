"""公式语法校验与范围校验。"""

from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.utils import get_column_letter

from ._utils import open_workbook, parse_cell_range, validate_cell_ref
from .exceptions import ValidationError

logger = logging.getLogger(__name__)

_UNSAFE_FUNCS = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}
_FUNC_RE = re.compile(r"([A-Z]+)\s*\(")
_CELL_REF_RE = re.compile(r"[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?")


def validate_formula(formula: str) -> tuple[bool, str]:
    """校验公式语法与安全性（括号配平、危险函数）。"""
    if not formula.startswith("="):
        return False, "Formula must start with '='"
    body = formula[1:]
    depth = 0
    for ch in body:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if depth < 0:
            return False, "Unmatched closing parenthesis"
    if depth > 0:
        return False, "Unclosed parenthesis"
    for func in _FUNC_RE.findall(body):
        if func in _UNSAFE_FUNCS:
            return False, f"Unsafe function: {func}"
    return True, "Formula is valid"


def validate_formula_in_cell_operation(
    filepath: str, sheet_name: str, cell: str, formula: str
) -> dict[str, Any]:
    """校验公式语法，并与单元格现有公式比较。"""
    try:
        if not validate_cell_ref(cell):
            raise ValidationError(f"Invalid cell reference: {cell}")
        is_valid, message = validate_formula(formula)
        if not is_valid:
            raise ValidationError(f"Invalid formula syntax: {message}")
        # 校验公式中的单元格引用
        for ref in _CELL_REF_RE.findall(formula):
            parts = ref.split(":") if ":" in ref else [ref]
            for p in parts:
                if not validate_cell_ref(p):
                    raise ValidationError(f"Invalid cell reference in formula: {ref}")

        with open_workbook(filepath) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet {sheet_name!r} not found")
            current = wb[sheet_name][cell].value

        normalized = formula if formula.startswith("=") else f"={formula}"
        if isinstance(current, str) and current.startswith("="):
            matches = current == normalized
            return {
                "message": "Formula is valid and matches cell content" if matches else "Formula is valid but doesn't match cell content",
                "valid": True,
                "matches": matches,
                "cell": cell,
                "provided_formula": normalized,
                "current_formula": current,
            }
        return {
            "message": "Formula is valid but cell contains no formula",
            "valid": True,
            "matches": False,
            "cell": cell,
            "provided_formula": normalized,
            "current_content": str(current) if current is not None else "",
        }
    except ValidationError:
        raise
    except Exception as e:
        logger.error("Failed to validate formula: %s", e)
        raise ValidationError(str(e)) from e


def validate_range_in_sheet_operation(
    filepath: str, sheet_name: str, start_cell: str, end_cell: str | None = None
) -> dict[str, Any]:
    """校验范围是否在工作表数据边界内，并返回数据范围信息。"""
    try:
        with open_workbook(filepath) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet {sheet_name!r} not found")
            ws = wb[sheet_name]
            data_max_row = ws.max_row
            data_max_col = ws.max_column
            sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
            if er is None:
                er = sr
            if ec is None:
                ec = sc
            if sr < 1 or sr > max(data_max_row, 1):
                raise ValidationError(f"Start row {sr} out of bounds (1-{data_max_row})")
            if sc < 1 or sc > max(data_max_col, 1):
                raise ValidationError(f"Start column {get_column_letter(sc)} out of bounds (A-{get_column_letter(data_max_col)})")
            if er < sr:
                raise ValidationError("End row cannot be before start row")
            if ec < sc:
                raise ValidationError("End column cannot be before start column")

            rng = start_cell if end_cell is None else f"{start_cell}:{end_cell}"
            data_rng = f"A1:{get_column_letter(data_max_col)}{data_max_row}" if data_max_col and data_max_row else "A1:A1"
            extends = er > data_max_row or ec > data_max_col
            return {
                "message": f"Range {rng!r} is valid. Sheet contains data in range {data_rng!r}",
                "valid": True,
                "range": rng,
                "data_range": data_rng,
                "extends_beyond_data": extends,
                "data_dimensions": {
                    "max_row": data_max_row,
                    "max_col": data_max_col,
                    "max_col_letter": get_column_letter(data_max_col) if data_max_col else "A",
                },
            }
    except ValidationError:
        raise
    except Exception as e:
        logger.error("Failed to validate range: %s", e)
        raise ValidationError(str(e)) from e


def validate_range_bounds(ws, sr: int, sc: int, er: int | None = None, ec: int | None = None) -> tuple[bool, str]:
    """校验范围是否在工作表边界内（保留原 API）。"""
    max_row = ws.max_row
    max_col = ws.max_column
    if sr < 1 or sr > max_row:
        return False, f"Start row {sr} out of bounds (1-{max_row})"
    if sc < 1 or sc > max_col:
        return False, f"Start column {get_column_letter(sc)} out of bounds (A-{get_column_letter(max_col)})"
    if er is not None and ec is not None:
        if er < sr:
            return False, "End row cannot be before start row"
        if ec < sc:
            return False, "End column cannot be before start column"
        if er > max_row:
            return False, f"End row {er} out of bounds (1-{max_row})"
        if ec > max_col:
            return False, f"End column {get_column_letter(ec)} out of bounds (A-{get_column_letter(max_col)})"
    return True, "Range is valid"
