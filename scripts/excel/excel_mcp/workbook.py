"""工作簿级操作：创建、元数据。"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from ._utils import col_index_to_letter, resolve_path
from .exceptions import WorkbookError

logger = logging.getLogger(__name__)


def create_workbook(filepath: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """创建新工作簿（含一个默认工作表）。父目录自动创建。"""
    try:
        path = Path(resolve_path(filepath))
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb["Sheet"].title = sheet_name
        else:
            wb.create_sheet(sheet_name)
        wb.save(str(path))
        return {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "path": str(path),
        }
    except Exception as e:
        logger.error("Failed to create workbook: %s", e)
        raise WorkbookError(f"Failed to create workbook: {e}") from e


def create_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    """在已有工作簿中新建工作表，已存在则报错。"""
    from ._utils import edit_workbook
    from .exceptions import SheetError
    try:
        with edit_workbook(filepath) as wb:
            if sheet_name in wb.sheetnames:
                raise SheetError(f"Sheet {sheet_name!r} already exists")
            wb.create_sheet(sheet_name)
        return {"message": f"Sheet {sheet_name!r} created successfully"}
    except (SheetError, WorkbookError):
        raise
    except Exception as e:
        logger.error("Failed to create sheet: %s", e)
        raise WorkbookError(str(e)) from e


def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """获取工作簿元数据：文件名、工作表列表、大小、修改时间，可选各表使用范围。"""
    from ._utils import open_workbook
    from .exceptions import NotFoundError
    try:
        path = Path(resolve_path(filepath))
        if not path.exists():
            raise NotFoundError(f"File not found: {filepath}")
        stat = path.stat()
        info: dict[str, Any] = {
            "filename": path.name,
            "path": str(path),
            "sheets": [],
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        }
        with open_workbook(str(path)) as wb:
            info["sheets"] = list(wb.sheetnames)
            if include_ranges:
                ranges: dict[str, str] = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    if ws.max_row and ws.max_column:
                        ranges[name] = f"A1:{col_index_to_letter(ws.max_column)}{ws.max_row}"
                info["used_ranges"] = ranges
        return info
    except (NotFoundError, WorkbookError):
        raise
    except Exception as e:
        logger.error("Failed to get workbook info: %s", e)
        raise WorkbookError(str(e)) from e
