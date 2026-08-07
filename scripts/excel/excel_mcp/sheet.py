"""工作表与范围级操作：复制/删除/重命名工作表、合并、范围复制/删除、行列增删。"""

from __future__ import annotations

import logging
from copy import copy
from typing import Any

from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from ._utils import (
    edit_workbook,
    open_workbook,
    parse_cell_range,
    parse_cell_ref,
    range_to_str,
    require_sheet,
    validate_cell_ref,
)
from .exceptions import SheetError, ValidationError

logger = logging.getLogger(__name__)


# ────────────────────────────── 工作表 ──────────────────────────────

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> dict[str, Any]:
    try:
        with edit_workbook(filepath) as wb:
            if source_sheet not in wb.sheetnames:
                raise SheetError(f"Source sheet {source_sheet!r} not found")
            if target_sheet in wb.sheetnames:
                raise SheetError(f"Target sheet {target_sheet!r} already exists")
            target = wb.copy_worksheet(wb[source_sheet])
            target.title = target_sheet
        return {"message": f"Sheet {source_sheet!r} copied to {target_sheet!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to copy sheet: %s", e)
        raise SheetError(str(e)) from e


def delete_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    try:
        with edit_workbook(filepath) as wb:
            if sheet_name not in wb.sheetnames:
                raise SheetError(f"Sheet {sheet_name!r} not found")
            if len(wb.sheetnames) == 1:
                raise SheetError("Cannot delete the only sheet in workbook")
            del wb[sheet_name]
        return {"message": f"Sheet {sheet_name!r} deleted"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to delete sheet: %s", e)
        raise SheetError(str(e)) from e


def rename_sheet(filepath: str, old_name: str, new_name: str) -> dict[str, Any]:
    try:
        with edit_workbook(filepath) as wb:
            if old_name not in wb.sheetnames:
                raise SheetError(f"Sheet {old_name!r} not found")
            if new_name in wb.sheetnames:
                raise SheetError(f"Sheet {new_name!r} already exists")
            wb[old_name].title = new_name
        return {"message": f"Sheet renamed from {old_name!r} to {new_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to rename sheet: %s", e)
        raise SheetError(str(e)) from e


# ────────────────────────────── 合并 ──────────────────────────────

def merge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    try:
        sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
        if er is None or ec is None:
            raise SheetError("Both start and end cells must be specified for merging")
        rng = range_to_str(sr, sc, er, ec)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            ws.merge_cells(rng)
        return {"message": f"Range {rng!r} merged in sheet {sheet_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to merge range: %s", e)
        raise SheetError(str(e)) from e


def unmerge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    try:
        sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
        if er is None or ec is None:
            raise SheetError("Both start and end cells must be specified for unmerging")
        rng = range_to_str(sr, sc, er, ec)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            merged = {str(m).upper() for m in ws.merged_cells.ranges}
            if rng.upper() not in merged:
                raise SheetError(f"Range {rng!r} is not merged")
            ws.unmerge_cells(rng)
        return {"message": f"Range {rng!r} unmerged successfully"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to unmerge range: %s", e)
        raise SheetError(str(e)) from e


def get_merged_ranges(filepath: str, sheet_name: str) -> list[str]:
    try:
        with open_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            return [str(m) for m in ws.merged_cells.ranges]
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to get merged cells: %s", e)
        raise SheetError(str(e)) from e


# ────────────────────────────── 范围复制/删除 ──────────────────────────────

def _copy_cell_style(src, dst) -> None:
    """安全复制单元格样式，忽略不支持属性。"""
    try:
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
    except Exception:
        dst.value = src.value


def copy_range_operation(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: str | None = None,
) -> dict[str, Any]:
    try:
        sr, sc, er, ec = parse_cell_range(source_start, source_end)
        if er is None or ec is None:
            er, ec = sr, sc
        tr, tc = parse_cell_ref(target_start)
        with edit_workbook(filepath) as wb:
            src_ws = require_sheet(wb, sheet_name)
            dst_ws = require_sheet(wb, target_sheet) if target_sheet else src_ws
            row_off = tr - sr
            col_off = tc - sc
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    _copy_cell_style(src_ws.cell(row=r, column=c), dst_ws.cell(row=r + row_off, column=c + col_off))
        return {"message": "Range copied successfully"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to copy range: %s", e)
        raise SheetError(str(e)) from e


def _clear_range(ws, start_cell: str, end_cell: str | None = None) -> None:
    sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
    if er is None or ec is None:
        er, ec = sr, sc
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.number_format = "General"
            cell.alignment = None


def delete_range_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
    shift_direction: str = "up",
) -> dict[str, Any]:
    try:
        if shift_direction not in ("up", "left"):
            raise ValidationError(f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'")
        sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
        if er is None or ec is None:
            er, ec = sr, sc
        rng = range_to_str(sr, sc, er, ec)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            _clear_range(ws, start_cell, end_cell)
            if shift_direction == "up":
                ws.delete_rows(sr, er - sr + 1)
            else:
                ws.delete_cols(sc, ec - sc + 1)
        return {"message": f"Range {rng} deleted successfully"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to delete range: %s", e)
        raise SheetError(str(e)) from e


# ────────────────────────────── 行列增删 ──────────────────────────────

def _validate_count(start: int, count: int, *, is_row: bool) -> None:
    if start < 1:
        raise ValidationError(f"Start {'row' if is_row else 'column'} must be 1 or greater")
    if count < 1:
        raise ValidationError("Count must be 1 or greater")


def insert_row(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> dict[str, Any]:
    try:
        _validate_count(start_row, count, is_row=True)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            ws.insert_rows(start_row, count)
        return {"message": f"Inserted {count} row(s) starting at row {start_row} in sheet {sheet_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to insert rows: %s", e)
        raise SheetError(str(e)) from e


def insert_cols(filepath: str, sheet_name: str, start_col: int, count: int = 1) -> dict[str, Any]:
    try:
        _validate_count(start_col, count, is_row=False)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            ws.insert_cols(start_col, count)
        return {"message": f"Inserted {count} column(s) starting at column {start_col} in sheet {sheet_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to insert columns: %s", e)
        raise SheetError(str(e)) from e


def delete_rows(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> dict[str, Any]:
    try:
        _validate_count(start_row, count, is_row=True)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            if start_row > ws.max_row:
                raise ValidationError(f"Start row {start_row} exceeds worksheet bounds (max row: {ws.max_row})")
            ws.delete_rows(start_row, count)
        return {"message": f"Deleted {count} row(s) starting at row {start_row} in sheet {sheet_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to delete rows: %s", e)
        raise SheetError(str(e)) from e


def delete_cols(filepath: str, sheet_name: str, start_col: int, count: int = 1) -> dict[str, Any]:
    try:
        _validate_count(start_col, count, is_row=False)
        with edit_workbook(filepath) as wb:
            ws = require_sheet(wb, sheet_name)
            if start_col > ws.max_column:
                raise ValidationError(f"Start column {start_col} exceeds worksheet bounds (max column: {ws.max_column})")
            ws.delete_cols(start_col, count)
        return {"message": f"Deleted {count} column(s) starting at column {start_col} in sheet {sheet_name!r}"}
    except (SheetError, ValidationError):
        raise
    except Exception as e:
        logger.error("Failed to delete columns: %s", e)
        raise SheetError(str(e)) from e
