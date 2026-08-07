"""单元格格式化：字体、填充、边框、对齐、数字格式、合并、保护、条件格式。"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule, IconSetRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side

from ._utils import edit_workbook, get_or_create_workbook, parse_cell_range, require_sheet, validate_cell_ref
from .exceptions import FormattingError, ValidationError

logger = logging.getLogger(__name__)


def _normalize_color(color: str) -> str:
    """确保 8 位 ARGB（FF 前缀）。"""
    color = color.lstrip("#")
    if len(color) == 6:
        return f"FF{color}"
    if len(color) == 8:
        return color
    raise FormattingError(f"Invalid color: {color!r}")


_CONDITIONAL_RULES = {
    "color_scale": ColorScaleRule,
    "data_bar": DataBarRule,
    "icon_set": IconSetRule,
    "formula": FormulaRule,
    "cell_is": CellIsRule,
}


def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    border_style: str | None = None,
    border_color: str | None = None,
    number_format: str | None = None,
    alignment: str | None = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: dict[str, Any] | None = None,
    conditional_format: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """对范围应用格式化。"""
    try:
        if not validate_cell_ref(start_cell):
            raise ValidationError(f"Invalid start cell reference: {start_cell}")
        if end_cell and not validate_cell_ref(end_cell):
            raise ValidationError(f"Invalid end cell reference: {end_cell}")

        sr, sc, er, ec = parse_cell_range(start_cell, end_cell)
        if er is None:
            er = sr
        if ec is None:
            ec = sc

        # 字体
        font_args: dict[str, Any] = {"bold": bold, "italic": italic, "underline": "single" if underline else None}
        if font_size is not None:
            font_args["size"] = font_size
        if font_color is not None:
            font_args["color"] = Color(rgb=_normalize_color(font_color))
        font = Font(**font_args)

        # 填充
        fill = None
        if bg_color is not None:
            argb = _normalize_color(bg_color)
            fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")

        # 边框
        border = None
        if border_style is not None:
            bc = _normalize_color(border_color) if border_color else "FF000000"
            side = Side(style=border_style, color=Color(rgb=bc))
            border = Border(left=side, right=side, top=side, bottom=side)

        # 对齐
        align = None
        if alignment is not None or wrap_text:
            align = Alignment(horizontal=alignment, vertical="center", wrap_text=wrap_text)

        # 保护
        protect = Protection(**protection) if protection is not None else None

        wb = get_or_create_workbook(filepath)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet {sheet_name!r} not found")
            ws = wb[sheet_name]
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = font
                    if fill is not None:
                        cell.fill = fill
                    if border is not None:
                        cell.border = border
                    if align is not None:
                        cell.alignment = align
                    if protect is not None:
                        cell.protection = protect
                    if number_format is not None:
                        cell.number_format = number_format

            if merge_cells and end_cell:
                ws.merge_cells(f"{start_cell}:{end_cell}")

            if conditional_format is not None:
                _apply_conditional_format(ws, start_cell, end_cell, conditional_format)

            wb.save(filepath)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        rng = f"{start_cell}:{end_cell}" if end_cell else start_cell
        return {"message": f"Applied formatting to range {rng}", "range": rng}
    except (ValidationError, FormattingError):
        raise
    except Exception as e:
        logger.error("Failed to apply formatting: %s", e)
        raise FormattingError(str(e)) from e


def _apply_conditional_format(ws, start_cell: str, end_cell: str | None, cf: dict[str, Any]) -> None:
    rule_type = cf.get("type")
    if not rule_type:
        raise FormattingError("Conditional format type not specified")
    params = dict(cf.get("params", {}))
    rng = f"{start_cell}:{end_cell}" if end_cell else start_cell

    if rule_type == "cell_is" and isinstance(params.get("fill"), dict):
        fill_params = params["fill"]
        fg = fill_params.get("fgColor", "FFC7CE")
        fg = _normalize_color(fg)
        params["fill"] = PatternFill(start_color=fg, end_color=fg, fill_type="solid")

    rule_cls = _CONDITIONAL_RULES.get(rule_type)
    if rule_cls is None:
        raise FormattingError(f"Invalid conditional format type: {rule_type}")
    try:
        rule = rule_cls(**params)
    except Exception as e:
        raise FormattingError(f"Failed to build conditional rule: {e}") from e
    ws.conditional_formatting.add(rng, rule)
